from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

wb = load_workbook('C:\\Users\\aivan\\Desktop\\Python_programs\\Openpyxl\\Grades.xlsx')
ws = wb.active
ws.title = "Data"
print(ws['A1'].value)
#Cambiar valor
ws['A2'].value = "Test"
#Esconder columnas?
ws.merge_cells("A1:D2")

#Insertar filas
ws.insert_rows(7)

#Insertar columnas
ws.insert_cols(2)

#Borras filas
ws.delete_rows(7)

#Borras columnas
ws.delete_cols(2)

#Mover columnas
ws.move_range("C1:D11", rows=2, cols=2)

#Recorrer sheet
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

for i in range(1,11):
    print(2 * i)


#Append to excel
#ws.append(['Valor 1','Valor 2', 'Valor 3','Valor 4'])
#Crear sheet
#wb.create_sheet("Test")
#Imprimir los nombres de las sheets
#print(wb.sheetnames)
#Guardar después de hacer cambios   
#wb.save('C:\\Users\\aivan\\Desktop\\Python_programs\\Openpyxl\\Grades.xlsx')
